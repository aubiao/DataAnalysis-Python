import os
import requests
import pandas as pd
import xlsxwriter
import openpyxl
from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows


def savepics(img_urls, titles):
    for i in range(len(img_urls)):
        img_url = img_urls[i]
        title = titles[i]
        img_data = requests.get(img_url).content

        with open(str(title)+'.jpg', 'wb') as f:
            f.write(img_data)


def csvToExcel1(books_data):
    titles = books_data['titles']
    authors = books_data['authors']
    ratings = books_data['ratings']
    details = books_data['details']

    workbook = xlsxwriter.Workbook('Books.xlsx')
    worksheet = workbook.add_worksheet('豆瓣新书')

    nums = len(titles)

    worksheet.write(0, 0, '图书封面')
    worksheet.write(0, 1, '图书标题')
    worksheet.write(0, 2, '图书作者')
    worksheet.write(0, 3, '图书评价')
    worksheet.write(0, 4, '图书细节')

    worksheet.set_column('A:A', 40)
    worksheet.set_column('B:B', 20)
    worksheet.set_column('C:C', 20)
    worksheet.set_column('D:D', 10)
    worksheet.set_column('E:E', 150)

    for i in range(1, nums):
        worksheet.insert_image(i, 0, titles[i]+'.jpg')
        worksheet.write(i, 1, titles[i])
        worksheet.write(i, 2, authors[i])
        worksheet.write(i, 3, str(ratings[i]))
        worksheet.write(i, 4, details[i])

    workbook.close()


def csvToExcel2(books_data):
    df1 = pd.DataFrame(books_data)

    data = {'代号': ['A', 'B', 'C', 'D'], '身高': [
        178, 177, 180, 175], '体重': [65, 70, 64, 67]}
    df2 = pd.DataFrame(data)

    writer = pd.ExcelWriter('pandas_moresheet.xlsx', engine='xlsxwriter')

    df1.to_excel(writer, sheet_name='豆瓣图书')
    df2.to_excel(writer, sheet_name='体测数据')

    writer.save()

    df = pd.read_excel('pandas_moresheet.xlsx', sheet_name='体测数据')
    print(df)


def csvToExcel3(books_data):
    df1 = pd.DataFrame(books_data)

    data = {'代号': ['A', 'B', 'C', 'D'], '身高': [
        178, 177, 180, 175], '体重': [65, 70, 64, 67]}
    df2 = pd.DataFrame(data)

    wb = Workbook()
    ws1 = wb.create_sheet("豆瓣图书", 0)
    ws2 = wb.create_sheet("体测数据", 1)

    nums = len(books_data)
    titles = books_data['titles']

    for i in range(1, nums):
        img = openpyxl.drawing.image.Image(titles[i]+'.jpg')
        ws1.add_image(img, 'B'+str(i + 2))

    for r in dataframe_to_rows(df1, index=True, header=True):
        ws1.append(r)

    for r in dataframe_to_rows(df2, index=True, header=True):
        ws2.append(r)

    wb.save("pandas_openpyxl.xlsx")

    df = pd.read_excel('pandas_openpyxl.xlsx')
    print(df)


if __name__ == '__main__':
    if 'ExcelData' not in os.listdir():
        os.mkdir('ExcelData')
    os.chdir('ExcelData')

    # books_data = pd.read_csv(
    #     '..\doubanbook.csv', usecols=['img_urls', 'titles'])
    # img_urls = books_data['img_urls']
    # titles = books_data['titles']
    # savepics(img_urls, titles)

    books_data = pd.read_csv('..\doubanbook.csv', usecols=[
        'titles', 'authors', 'ratings', 'details'
    ], na_values='NULL')

    # csvToExcel1(books_data)
    # csvToExcel2(books_data)
    csvToExcel3(books_data)
